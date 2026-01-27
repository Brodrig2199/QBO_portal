import os
import secrets
import base64
import io
import html
from datetime import datetime, timedelta, timezone
from functools import wraps

import requests
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file

from token_store import init_db, save_tokens
from qbo_client import (
    get_valid_access_token,
    get_customers,
    get_accounts,
    get_profit_and_loss_detail,
    get_vat_tax_detail,
    parse_report_to_table,
    get_vendors,
    get_vendor_detail,
    extract_vendor_otro,
)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-change-me")

try:
    init_db()
except Exception as e:
    print("DB init skipped:", e)

LOGIN_USER = os.environ.get("LOGIN_USER", "admin")
LOGIN_PASS = os.environ.get("LOGIN_PASS", "admin123")

REPORT_TYPES = [
    {"id": "profit_and_loss_detail", "name": "Detalle de Pérdidas y Ganancias", "qbo": "ProfitAndLossDetail"},
    {"id": "vat_tax_detail", "name": "VAT - Detalle de Impuestos", "qbo": "TaxDetail"},
]


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def parse_date(date_str: str) -> str:
    datetime.strptime(date_str, "%Y-%m-%d")
    return date_str


def fetch_qbo_report(report_type: str, start_date: str, end_date: str, client_id: str, excluded_accounts: list[str]):
    access_token, realm_id = get_valid_access_token()

    if report_type == "profit_and_loss_detail":
        report_json = get_profit_and_loss_detail(
            access_token=access_token,
            realm_id=realm_id,
            start_date=start_date,
            end_date=end_date,
            accounting_method="Accrual",
            customer_id=None if client_id == "all" else client_id
        )
        table = parse_report_to_table(report_json)

        return {"meta": {"report_type": report_type, "qbo_report_name": "ProfitAndLossDetail",
                         "start_date": start_date, "end_date": end_date, "client_id": client_id,
                         "accounting_method": "Accrual",
                         "excluded_accounts": excluded_accounts},
                "table": table, "raw": report_json}

    if report_type == "vat_tax_detail":
        report_json = get_vat_tax_detail(access_token, realm_id, start_date, end_date)
        table = parse_report_to_table(report_json)

        return {"meta": {"report_type": report_type, "qbo_report_name": "TaxDetail",
                         "start_date": start_date, "end_date": end_date, "client_id": client_id,
                         "excluded_accounts": excluded_accounts},
                "table": table, "raw": report_json}

    raise RuntimeError(f"Tipo de reporte inválido: {report_type}")


@app.get("/")
def home():
    return redirect(url_for("reports")) if session.get("logged_in") else redirect(url_for("login"))


@app.get("/login")
def login():
    return render_template("login.html")


@app.post("/login")
def login_post():
    username = request.form.get("username", "")
    password = request.form.get("password", "")
    if username == LOGIN_USER and password == LOGIN_PASS:
        session["logged_in"] = True
        session["username"] = username
        return redirect(url_for("reports"))
    flash("Usuario o contraseña incorrectos.")
    return redirect(url_for("login"))


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.get("/connect")
def connect():
    client_id = os.environ.get("QBO_CLIENT_ID", "")
    redirect_uri = os.environ.get("QBO_REDIRECT_URI", "")
    if not client_id or not redirect_uri:
        return "Faltan QBO_CLIENT_ID o QBO_REDIRECT_URI en env vars", 500

    authorize_url = "https://appcenter.intuit.com/connect/oauth2"
    scope = "com.intuit.quickbooks.accounting"

    state = secrets.token_urlsafe(24)
    session["oauth_state"] = state
    session["after_auth"] = request.args.get("next") or url_for("reports")

    from urllib.parse import urlencode
    params = {
        "client_id": client_id,
        "response_type": "code",
        "scope": scope,
        "redirect_uri": redirect_uri,
        "state": state,
    }
    return redirect(f"{authorize_url}?{urlencode(params)}")


@app.get("/callback")
def callback():
    code = request.args.get("code")
    realm_id = request.args.get("realmId")
    state = request.args.get("state")
    err = request.args.get("error")

    if err:
        return f"Autorización falló: {request.args.get('error_description', err)}", 400

    saved_state = session.get("oauth_state")
    if not saved_state or state != saved_state:
        return "State inválido. Reintenta /connect.", 400

    if not code or not realm_id:
        return "Faltan parámetros code o realmId.", 400

    client_id = os.environ.get("QBO_CLIENT_ID", "")
    client_secret = os.environ.get("QBO_CLIENT_SECRET", "")
    redirect_uri = os.environ.get("QBO_REDIRECT_URI", "")
    if not client_id or not client_secret or not redirect_uri:
        return "Faltan env vars QBO_CLIENT_ID/SECRET/REDIRECT_URI", 500

    token_url = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"
    basic = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()

    headers = {
        "Authorization": f"Basic {basic}",
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    data = {"grant_type": "authorization_code", "code": code, "redirect_uri": redirect_uri}

    r = requests.post(token_url, headers=headers, data=data, timeout=30)
    if r.status_code >= 400:
        return f"Token exchange failed ({r.status_code}): {r.text}", 400

    payload = r.json()
    access_token = payload.get("access_token")
    refresh_token = payload.get("refresh_token")
    expires_in = int(payload.get("expires_in", 3600))

    expires_at = datetime.now(timezone.utc) + timedelta(seconds=expires_in)
    save_tokens(realm_id=realm_id, access_token=access_token, refresh_token=refresh_token, access_expires_at=expires_at)

    session.pop("oauth_state", None)
    flash("QuickBooks conectado ✅")
    return redirect(session.pop("after_auth", url_for("reports")))


@app.get("/reports")
@login_required
def reports():
    try:
        access_token, realm_id = get_valid_access_token()
        clients = [{"id": "all", "name": "Todos los clientes"}] + get_customers(access_token, realm_id)
        accounts = get_accounts(access_token, realm_id)
        return render_template("reports.html", clients=clients, accounts=accounts, report_types=REPORT_TYPES)
    except Exception as e:
        print("REPORTS ERROR ->", repr(e))
        flash(f"QuickBooks no conectado o error: {e}. Ve a /connect.")
        return render_template("reports.html", clients=[{"id": "all", "name": "Todos los clientes"}], accounts=[], report_types=REPORT_TYPES)


@app.post("/run-report")
@login_required
def run_report():
    try:
        report_type = request.form.get("report_type", "")
        start_date = parse_date(request.form.get("start_date", ""))
        end_date = parse_date(request.form.get("end_date", ""))
        client_id = request.form.get("client_id", "all")
        excluded_accounts = request.form.getlist("excluded_accounts")

        print("RUN REPORT -> report_type:", report_type, "start:", start_date, "end:", end_date, "client:", client_id)

        data = fetch_qbo_report(report_type, start_date, end_date, client_id, excluded_accounts)

        # Guardar meta para download
        session["last_report_meta"] = data["meta"]

        return render_template("results.html", data=data)

    except Exception as e:
        print("RUN REPORT ERROR ->", repr(e))
        flash(f"Error generando reporte: {e}")
        return redirect(url_for("reports"))


@app.get("/download/qbo/report.xlsx")
@login_required
def download_qbo_report_xlsx():
    meta = session.get("last_report_meta")
    if not meta:
        flash("No hay parámetros del reporte. Genera uno primero.")
        return redirect(url_for("reports"))

    access_token, realm_id = get_valid_access_token()

    # 🔹 Re-descargar el reporte original de QuickBooks (preview completo)
    if meta["report_type"] == "profit_and_loss_detail":
        report_json = get_profit_and_loss_detail(
            access_token=access_token,
            realm_id=realm_id,
            start_date=meta["start_date"],
            end_date=meta["end_date"],
            accounting_method="Accrual",
            customer_id=None if meta.get("client_id") in (None, "", "all") else meta["client_id"],
        )
        sheet_title = "Profit & Loss Detail"
        filename = f"QBO_ProfitAndLossDetail_{meta['start_date']}_{meta['end_date']}.xlsx"

    elif meta["report_type"] == "vat_tax_detail":
        report_json = get_vat_tax_detail(
            access_token=access_token,
            realm_id=realm_id,
            start_date=meta["start_date"],
            end_date=meta["end_date"],
        )
        sheet_title = "VAT Tax Detail"
        filename = f"QBO_TaxDetail_{meta['start_date']}_{meta['end_date']}.xlsx"

    else:
        flash("Tipo de reporte no soportado.")
        return redirect(url_for("reports"))

    table = parse_report_to_table(report_json)

    # --- Crear Excel genérico (tal cual QuickBooks) ---
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # headers
    ws.append(table["columns"])
    for c in range(1, len(table["columns"]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    # rows
    for r in table["rows"]:
        ws.append(r["cells"])

    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/download/informe43.xlsx")
@login_required
def download_informe43_xlsx():
    meta = session.get("last_report_meta")
    if not meta:
        flash("No hay parámetros del reporte. Genera uno primero.")
        return redirect(url_for("reports"))

    # INFORME 43 basado en P&L DETAIL
    if meta.get("report_type") != "profit_and_loss_detail":
        flash("El INFORME 43 se genera desde Detalle de Pérdidas y Ganancias.")
        return redirect(url_for("reports"))

    access_token, realm_id = get_valid_access_token()

    report_json = get_profit_and_loss_detail(
        access_token=access_token,
        realm_id=realm_id,
        start_date=meta["start_date"],
        end_date=meta["end_date"],
        accounting_method="Accrual",
        customer_id=None if meta.get("client_id") in (None, "", "all") else meta["client_id"],
    )

    table = parse_report_to_table(report_json)

    # -------------------------
    # Helpers
    # -------------------------
    import re, io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    cols = [(c or "").strip().lower() for c in (table.get("columns") or [])]

    def find_col_contains(*keys):
        keys = [k.lower() for k in keys]
        for i, c in enumerate(cols):
            for k in keys:
                if k in c:
                    return i
        return None

    def cell(row, idx):
        if idx is None:
            return ""
        cells = row.get("cells") or []
        if idx < 0 or idx >= len(cells):
            return ""
        return (cells[idx] or "").strip()

    def to_float(x):
        s = (x or "").strip()
        if not s:
            return 0.0
        try:
            return float(s.replace(",", ""))
        except:
            return 0.0

    def to_yyyymmdd(s):
        s = (s or "").strip()
        if not s:
            return ""
        for f in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, f).strftime("%Y%m%d")
            except:
                pass
        if len(s) == 8 and s.isdigit():
            return s
        return ""

    def is_panama_cedula(ruc: str) -> bool:
        s = (ruc or "").strip().upper()
        return bool(re.match(r'^\d{1,2}-\d{1,6}-\d{1,6}$', s))

    def infer_tipo_persona(tipo_from_name: str, ruc: str) -> str:
        t = (tipo_from_name or "").strip().upper()
        if t in ("N", "J", "E"):
            return t

        r = (ruc or "").strip().upper()
        if is_panama_cedula(r):
            return "N"
        if r.startswith("E") or "PASAPORTE" in r or "PASS" in r:
            return "E"

        digits = re.sub(r"\D", "", r)
        if len(digits) >= 10:
            return "J"

        return ""

    def normalize_factura(factura_raw: str, seq_num: int) -> str:
        f = (factura_raw or "").strip()
        return f if f else f"F-{seq_num}"

    def parse_vendor(name):
        """
        Soporta:
        1) NOMBRE/TIPO/RUC/DV   ej: BANCO GENERAL/2/280-134-61098/2
        2) NOMBRE/TIPO         ej: AMAZON/3
        """
        raw = (name or "").strip()
        if not raw:
            return ("", "", "", "")

        tipo_map = {"1": "N", "2": "J", "3": "E"}

        m = re.match(r'^\s*(.+?)\s*/\s*([123])\s*/\s*([^/]+)\s*/\s*([^/]+)\s*$', raw)
        if m:
            return (tipo_map.get(m.group(2).strip(), ""), m.group(3).strip(), m.group(4).strip(), m.group(1).strip())

        m2 = re.match(r'^\s*(.+?)\s*/\s*([123])\s*$', raw)
        if m2:
            return (tipo_map.get(m2.group(2).strip(), ""), "", "", m2.group(1).strip())

        return ("", "", "", raw.replace("/", " ").strip())
    

    def norm_key(s: str) -> str:
        s = html.unescape((s or "").strip())
        s = re.sub(r"\s+", " ", s)
        return s.lower().strip()

    def extract_ruc_dv_from_display(display_name: str):
        """
        From vendor DisplayName like:
        "ABOLU, S.A/2/16429-109-156121/61"
        returns ("16429-109-156121", "61")
        """
        dn = html.unescape((display_name or "").strip())
        m = re.match(r"^(.+?)/([123])/([^/]+)/([^/]+)\s*$", dn)
        if not m:
            return ("", "")
        ruc = (m.group(3) or "").strip()
        dv  = (m.group(4) or "").strip()
        return (ruc, dv)

    def parse_otros(notes_raw: str):
        """
        NOTES puede venir como:
        "2/1"
        "Concepto 2/1 algo"
        "2 / 1"
        Extrae el primer patrón num/num que encuentre.
        """
        s = (notes_raw or "").strip()
        if not s:
            return ("", "")

        m = re.search(r'(\d+)\s*/\s*(\d+)', s)
        if not m:
            return ("", "")

        return (m.group(1), m.group(2))

    # -------------------------
    # Column mapping (P&L Detail)
    # -------------------------
    idx_fecha   = find_col_contains("fecha", "date")
    idx_no      = find_col_contains("n.", "no", "nº", "numero", "number")
    idx_nombre  = find_col_contains("nombre", "name")
    idx_importe = find_col_contains("importe", "amount")
   


    # Cuenta contable (tu columna de QuickBooks)
    idx_cuenta_contable = find_col_contains("cuenta de división de artículo", "cuenta de division de articulo", "cuenta contable", "account")

    # -------------------------
    # 1) Sacar lista de vendors del reporte (NOMBRE limpio)
    # -------------------------
    vendor_names_needed = set()
    
    for r in (table.get("rows") or []):
        if r.get("is_header") or r.get("is_summary"):
            continue
        nombre_raw = cell(r, idx_nombre)
        if not nombre_raw:
            continue
        _, _, _, nombre_limpio = parse_vendor(nombre_raw)
        if nombre_limpio:
            vendor_names_needed.add(nombre_limpio.strip().lower())

    # -------------------------
    # 2) Construir mapa DisplayName->Id y luego Id->Other usando Batch
    # -------------------------
    from qbo_client import get_all_vendors_map, get_vendor_notes_by_ids

 
    vendors_map = get_all_vendors_map(access_token, realm_id) or {}
    # vendors_map esperado: {display_lower: id}

    # ✅ índice por RUC|DV (lo más confiable)
    rucdv_to_id = {}
    # ✅ índice por display exacto (por si el reporte trae el nombre completo)
    display_to_id = { norm_key(k): str(v) for k, v in vendors_map.items() }

    for disp_lower, vid in display_to_id.items():
        ruc, dv = extract_ruc_dv_from_display(disp_lower)
        if ruc and dv:
            rucdv_to_id[f"{ruc}|{dv}"] = str(vid)

    # ✅ ahora decidimos qué IDs buscar (de los vendors del reporte)
    ids_to_fetch = set()

    for r in (table.get("rows") or []):
        if r.get("is_header") or r.get("is_summary"):
            continue

        nombre_raw = cell(r, idx_nombre)
        if not nombre_raw:
            continue

        tipo_from_name, ruc_from_name, dv, nombre = parse_vendor(nombre_raw)

        # 1) intentar por RUC|DV
        vid = rucdv_to_id.get(f"{ruc_from_name}|{dv}")

        # 2) si no, intentar por display exacto (si el raw venía completo)
        if not vid:
            vid = display_to_id.get(norm_key(nombre_raw))

        # 3) si no, intentar por nombre limpio (último fallback)
        if not vid:
            vid = display_to_id.get(norm_key(nombre))

        if vid:
            ids_to_fetch.add(str(vid))

    ids_to_fetch = list(ids_to_fetch)

    # ✅ traer Notes por IDs
    vendor_notes_by_id = get_vendor_notes_by_ids(access_token, realm_id, ids_to_fetch) or {}
    print("DEBUG sample vendor notes:", list(vendor_notes_by_id.items())[:10])


    # -------------------------
    # 3) Construir filas INFORME 43
    # -------------------------
    rows_out = []
    seq = 1

    for r in (table.get("rows") or []):
        if r.get("is_header") or r.get("is_summary"):
            continue

        nombre_raw = cell(r, idx_nombre)
        if not nombre_raw:
            continue

        tipo_from_name, ruc_from_name, dv, nombre = parse_vendor(nombre_raw)

        # Tipo final N/J/E arreglado (con RUC)
        tipo = infer_tipo_persona(tipo_from_name, ruc_from_name)

        # Factura con fallback
        factura = normalize_factura(cell(r, idx_no), seq)
        seq += 1

        # Montos
        monto_balboas = to_float(cell(r, idx_importe))
        itbms_pagado = 0.00  # ✅ autocompletar ITBMS con cero en P&L

        # Cuenta contable
        cuenta_contable = cell(r, idx_cuenta_contable)

        # ✅ Vendor Notes -> Concepto/Compras (por nombre limpio)
        # ✅ Resolver vendor_id (RUC|DV primero)
        vid = rucdv_to_id.get(f"{ruc_from_name}|{dv}")

        if not vid:
            # fallback por nombre (por si el vendor no tiene /ruc/dv en display)
            vid = display_to_id.get(norm_key(nombre_raw)) or display_to_id.get(norm_key(nombre))

        notes_raw = vendor_notes_by_id.get(str(vid), "") if vid else ""
        concepto, compras = parse_otros(notes_raw)



        # ✅ Eliminar totales negativos (monto)
        if monto_balboas < 0:
            continue
        if not concepto and not compras:
         print("DEBUG EMPTY notes -> vid:", vid, "nombre:", nombre, "notes:", notes_raw)


        rows_out.append([
            tipo,                        # TIPO DE PERSONA
            ruc_from_name,               # RUC
            dv,                          # DV
            nombre,                      # NOMBRE O RAZON SOCIAL
            factura,                     # FACTURA
            to_yyyymmdd(cell(r, idx_fecha)),  # FECHA
            concepto,                    # CONCEPTO (Vendor->Otro: antes del /)
            compras,                     # COMPRAS (Vendor->Otro: después del /)
            monto_balboas,               # MONTO EN BALBOAS
            itbms_pagado,                # ITBMS PAGADO (0.00)
            cuenta_contable,             # CUENTA CONTABLE
        ])

    # -------------------------
    # Crear Excel (ligero para Render)
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "INFORME 43"

    headers = [
        "TIPO DE PERSONA",
        "RUC",
        "DV",
        "NOMBRE O RAZON SOCIAL",
        "FACTURA",
        "FECHA",
        "CONCEPTO",
        "COMPRAS DE BIENES Y SERVICIOS",
        "MONTO EN BALBOAS",
        "ITBMS PAGADO EN BALBOAS",
        "CUENTA CONTABLE",
    ]

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EFEFEF")

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "INFORME 43 - FORMATO A DILIGENCIAR"
    ws["A1"].font = Font(bold=True, size=13)

    ws.append([])
    ws.append([])

    # Header en fila 5
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data desde fila 6
    for r_i, rowvals in enumerate(rows_out, start=6):
        ws.append(rowvals)

    # Formatos
    for r_i in range(6, 6 + len(rows_out)):
        ws.cell(row=r_i, column=9).number_format = '#,##0.00'   # MONTO
        ws.cell(row=r_i, column=10).number_format = '#,##0.00'  # ITBMS
        ws.cell(row=r_i, column=2).number_format = '@'          # RUC texto
        ws.cell(row=r_i, column=3).number_format = '@'          # DV texto

    widths = [14, 20, 8, 35, 14, 12, 12, 22, 16, 18, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"INFORME43_{meta['start_date']}_{meta['end_date']}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/download/informe43_vat.xlsx")
@login_required
def download_informe43_vat_xlsx():
    meta = session.get("last_report_meta")
    if not meta:
        flash("No hay parámetros del reporte. Genera uno primero.")
        return redirect(url_for("reports"))

    if meta.get("report_type") != "vat_tax_detail":
        flash("Para este INFORME 43 (VAT) primero genera el reporte: VAT - Detalle de Impuestos.")
        return redirect(url_for("reports"))

    access_token, realm_id = get_valid_access_token()

    report_json = get_vat_tax_detail(
        access_token=access_token,
        realm_id=realm_id,
        start_date=meta["start_date"],
        end_date=meta["end_date"],
    )

    table = parse_report_to_table(report_json)

    # -------------------------
    # Helpers
    # -------------------------
    import re, io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    cols = [(c or "").strip().lower() for c in table.get("columns", [])]

    def find_col_contains(*keywords):
        # match por "contiene", para soportar títulos largos
        for i, c in enumerate(cols):
            for k in keywords:
                if k.lower() in c:
                    return i
        return None

    def cell(row, idx):
        if idx is None:
            return ""
        cells = row.get("cells") or []
        if idx < 0 or idx >= len(cells):
            return ""
        return (cells[idx] or "").strip()

    def to_float(x):
        try:
            return float(str(x).replace(",", "").strip())
        except:
            return 0.0

    def to_yyyymmdd(s):
        s = (s or "").strip()
        if not s:
            return ""
        for f in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, f).strftime("%Y%m%d")
            except:
                pass
        return ""

    def parse_vendor(name):
        """
        Soporta:
        - COMPLETO: NOMBRE/TIPO/RUC/DV   Ej: BANCO GENERAL/2/280-134-61098/2
        - PARCIAL: NOMBRE/TIPO          Ej: AMAZON/3
        """
        raw = (name or "").strip()
        if not raw:
            return ("", "", "", "")

        tipo_map = {"1": "N", "2": "J", "3": "E"}

        m = re.match(r'^\s*(.+?)\s*/\s*([123])\s*/\s*([^/]+)\s*/\s*([^/]+)\s*$', raw)
        if m:
            return (
                tipo_map.get(m.group(2).strip(), ""),
                m.group(3).strip(),
                m.group(4).strip(),
                m.group(1).strip()
            )

        m2 = re.match(r'^\s*(.+?)\s*/\s*([123])\s*$', raw)
        if m2:
            return (
                tipo_map.get(m2.group(2).strip(), ""),
                "",  # RUC
                "",  # DV
                m2.group(1).strip()
            )

        return ("", "", "", raw.replace("/", " ").strip())

   
    # -------------------------
    # MAP COLUMNAS VAT
    # -------------------------
    cols = [(c or "").strip().lower() for c in table.get("columns", [])]

    def find_col_contains(*keywords):
        for k in keywords:
            k = (k or "").strip().lower()
            for i, c in enumerate(cols):
                if k and k in c:
                    return i
        return None

    idx_fecha = find_col_contains("fecha", "date")
    idx_no = find_col_contains("n.", "no", "numero")
    idx_ruc_cliente = find_col_contains("ruc no. de cliente", "ruc cliente")
    idx_ruc_proveedor = find_col_contains("ruc no. de proveedor", "ruc proveedor")
    idx_nombre = find_col_contains("nombre", "name")

    # ✅ Base imponible = "Importe sujeto a impuestos"
    idx_base = find_col_contains("importe sujeto a impuestos", "importe sujeto", "taxable")

    # ✅ ITBMS = "Importe" (pero no el que dice sujeto)
    idx_itbms = None
    for i, c in enumerate(cols):
        c2 = (c or "").strip().lower()
        if (c2 == "importe" or c2.startswith("importe")) and ("sujeto" not in c2):
            idx_itbms = i
            break

    # ✅ Nombre del impuesto
    idx_tax_name = find_col_contains("nombre del impuesto", "tax name", "impuesto")

    # -------------------------
    # Helpers extra
    # -------------------------
    def to_float_safe(x):
        try:
            return float(str(x).replace(",", "").strip() or "0")
        except:
            return 0.0

    def is_no_tax(base, itbms, tax_name):
        # "no tiene impuesto" = ITBMS 0 o sin nombre de impuesto
        if itbms == 0 and (not (tax_name or "").strip()):
            return True
        if itbms == 0 and (tax_name or "").strip().lower() in ("exento", "0", "sin impuesto"):
            return True
        # puedes endurecerlo si quieres: base > 0 y itbms == 0
        return (itbms == 0)

    def is_negative_any(base, itbms):
        return (base < 0) or (itbms < 0)

    def normalize_factura(factura_raw: str, seq_num: int) -> str:
        f = (factura_raw or "").strip()
        return f if f else f"F-{seq_num}"

    def parse_vendor(name):
        """
        Soporta 2 formatos:
        1) COMPLETO: NOMBRE/TIPO/RUC/DV   (BANCO GENERAL/2/280-134-61098/2)
        2) PARCIAL:  NOMBRE/TIPO          (AMAZON/3)
        """
        import re
        raw = (name or "").strip()
        if not raw:
            return ("", "", "", "")

        tipo_map = {"1": "N", "2": "J", "3": "E"}

        m = re.match(r'^\s*(.+?)\s*/\s*([123])\s*/\s*([^/]+)\s*/\s*([^/]+)\s*$', raw)
        if m:
            return (
                tipo_map.get(m.group(2).strip(), ""),
                m.group(3).strip(),
                m.group(4).strip(),
                m.group(1).strip()
            )

        m2 = re.match(r'^\s*(.+?)\s*/\s*([123])\s*$', raw)
        if m2:
            return (
                tipo_map.get(m2.group(2).strip(), ""),
                "",  # ruc vacío
                "",  # dv vacío
                m2.group(1).strip()
            )

        return ("", "", "", raw.replace("/", " ").strip())

    def is_panama_cedula(ruc: str) -> bool:
        s = (ruc or "").strip().upper()
        return bool(re.match(r'^\d{1,2}-\d{1,6}-\d{1,6}$', s))

    def infer_tipo_persona(tipo_from_name: str, ruc: str, nombre: str = "") -> str:
        t = (tipo_from_name or "").strip().upper()
        if t in ("N", "J", "E"):
            return t

        r = (ruc or "").strip().upper()

        # ✅ si es tipo 8-NT-***-*** decidir por nombre
        if is_ruc_nt(r):
            return "J" if looks_like_company(nombre) else "N"

        if is_panama_cedula(r):
            return "N"

        if r.startswith("E") or "PASAPORTE" in r or "PASS" in r:
            return "E"

        digits = re.sub(r"\D", "", r)
        if len(digits) >= 10:
            return "J"

        return ""
    def looks_like_company(nombre: str) -> bool:
        n = (nombre or "").upper()

        # señales típicas de empresa
        company_tokens = [
            "S.A", "SA", "S. A", "INC", "CORP", "CORPORATION", "LLC", "LTD", "SRL",
            "S. DE R.L", "S DE RL", "S.A.S", "SAS", "CO.", "COMPANY",
            "FUNDACION", "ASOCIACION", "MINISTERIO", "UNIVERSIDAD", "HOSPITAL",
            "CLINICA", "COOPERATIVA", "IGLESIA", "BANCO"
        ]
        if any(t in n for t in company_tokens):
            return True

        # si tiene muchos símbolos típicos de razón social
        if "&" in n or "," in n:
            return True

        # si tiene más de 3 palabras, suele ser entidad (heurística)
        words = [w for w in re.split(r"\s+", n) if w]
        if len(words) >= 4:
            return True

        return False
    def is_ruc_nt(ruc: str) -> bool:
        # ejemplo: 8-NT-123-456
        s = (ruc or "").strip().upper()
        return bool(re.match(r'^\d{1,2}-[A-Z]{1,3}-\d{1,6}-\d{1,6}$', s))

    # -------------------------
    # Construir filas y separar Informe 5 / Informe 6
    # -------------------------
    informe5 = []  # con impuesto
    informe6 = []  # sin impuesto
    seq = 1

    for r in (table.get("rows") or []):
        if r.get("is_header") or r.get("is_summary"):
            continue

        nombre_raw = (cell(r, idx_nombre) if idx_nombre is not None else "").strip()
        if not nombre_raw:
            continue

        tipo_from_name, ruc_from_name, dv, nombre = parse_vendor(nombre_raw)

        ruc = (
            (cell(r, idx_ruc_proveedor) if idx_ruc_proveedor is not None else "").strip()
            or (cell(r, idx_ruc_cliente) if idx_ruc_cliente is not None else "").strip()
            or (ruc_from_name or "").strip()
        )

        tipo = infer_tipo_persona(tipo_from_name, ruc, nombre)

        factura = normalize_factura(cell(r, idx_no) if idx_no is not None else "", seq)
        seq += 1

        fecha_raw = (cell(r, idx_fecha) if idx_fecha is not None else "").strip()

        # primero intenta tu convertidor
        fecha_fmt = to_yyyymmdd(fecha_raw)

        # si por alguna razón no pudo, limpia todo lo que no sea número
        if not fecha_fmt:
            fecha_fmt = re.sub(r"\D", "", fecha_raw)

        # ✅ garantizar sin guiones (por si acaso)
        fecha_fmt = fecha_fmt.replace("-", "")

        base = to_float_safe(cell(r, idx_base) if idx_base is not None else "")
        itbms = to_float_safe(cell(r, idx_itbms) if idx_itbms is not None else "")

        tax_name = (cell(r, idx_tax_name) if idx_tax_name is not None else "").strip()
        tax_name_l = tax_name.lower()

        if "(ventas" in tax_name_l or " ventas)" in tax_name_l:
             continue


        # ✅ eliminar negativos
        if is_negative_any(base, itbms):
            continue

        # columnas del formato + extras solicitadas
        row_out = [
            tipo,               # TIPO DE PERSONA
            ruc,                # RUC
            dv,                 # DV
            nombre,             # NOMBRE O RAZON SOCIAL
            factura,            # FACTURA
            fecha_fmt,          # FECHA
            "",                 # CONCEPTO (lo llenas en P&L; aquí VAT lo dejas como pediste antes)
            "",                 # COMPRAS (igual)
            base,               # MONTO EN BALBOAS = IMPORTE SUJETO A IMPUESTOS
            itbms,              # ITBMS PAGADO EN BALBOAS = IMPORTE
            "",                 # ORIGEN INFORME (lo pongo luego)
            tax_name            # NOMBRE DEL IMPUESTO
        ]

        # Clasificar informe 5/6
        if is_no_tax(base, itbms, tax_name):
            # informe 6: solo sin impuesto
            row_out[-2] = "INFORME 6"
            informe6.append(row_out)
        else:
            row_out[-2] = "INFORME 5"
            informe5.append(row_out)

    # -------------------------
    # 1) Eliminar duplicados exactos dentro de cada informe
    # -------------------------
    def dedup_exact(rows):
        seen = set()
        out = []
        for row in rows:
            key = tuple(row)  # TODO igual
            if key in seen:
                continue
            seen.add(key)
            out.append(row)
        return out

    informe5 = dedup_exact(informe5)
    informe6 = dedup_exact(informe6)

    # -------------------------
    # 2) Si informe 6 repite EXACTAMENTE una fila de informe 5 -> eliminar de informe 6
    #    (comparación "todo igual" excepto el ORIGEN INFORME)
    # -------------------------
    def key_without_origen(row):
        # row[-2] es ORIGEN INFORME
        tmp = list(row)
        tmp[-2] = ""   # ignorar origen
        return tuple(tmp)

    keys_5 = set(key_without_origen(r) for r in informe5)

    informe6_filtrado = []
    for r in informe6:
        if key_without_origen(r) in keys_5:
            continue
        informe6_filtrado.append(r)

    informe6 = informe6_filtrado

    # -------------------------
    # Resultado final: informe 5 + informe 6
    # -------------------------
    rows_out = informe5 + informe6
    # -------------------------
    # ✅ Si las facturas quedan iguales -> agregar 'E' al final
    # (si hay 3 iguales: 1 queda igual, 2 = E, 3 = EE, etc.)
    # -------------------------
    seen_fact = {}
    for row in rows_out:
        factura = (row[4] or "").strip()   # FACTURA está en índice 4
        if not factura:
            continue

        cnt = seen_fact.get(factura, 0)
        if cnt > 0:
            row[4] = factura + ("E" * cnt)
        seen_fact[factura] = cnt + 1

    # -------------------------
    # Crear Excel
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "INFORME 43 (VAT)"

    headers = [
        "TIPO DE PERSONA",
        "RUC",
        "DV",
        "NOMBRE O RAZON SOCIAL",
        "FACTURA",
        "FECHA",
        "CONCEPTO",
        "COMPRAS DE BIENES Y SERVICIOS",
        "MONTO EN BALBOAS",
        "ITBMS PAGADO EN BALBOAS",
        "ORIGEN INFORME",
        "NOMBRE DEL IMPUESTO",
    ]

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EFEFEF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "INFORME 43 - FORMATO A DILIGENCIAR (VAT)"
    ws["A1"].font = Font(bold=True, size=13)

    ws.append([])
    ws.append([])

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = bold
        c.fill = fill
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for rr, rowvals in enumerate(rows_out, start=6):
        for cc, val in enumerate(rowvals, start=1):
            cellx = ws.cell(row=rr, column=cc, value=val)
            cellx.border = border

            # Texto para RUC y DV
            if cc in (2, 3):
                cellx.number_format = '@'

            # Formato moneda para MONTO e ITBMS
            if cc in (9, 10):
                cellx.number_format = '#,##0.00'

    widths = [16, 18, 6, 35, 14, 12, 18, 30, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"INFORME43_VAT_{meta['start_date']}_{meta['end_date']}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
